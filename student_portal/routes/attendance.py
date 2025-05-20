from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, make_response
from flask_login import login_required, current_user
from student_portal.models.db import get_db
from datetime import datetime, timedelta
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import calendar
import csv
from io import StringIO
import pdfkit
from jinja2 import Template
from io import BytesIO
from weasyprint import HTML

bp = Blueprint('attendance', __name__, url_prefix='/attendance')

@bp.route('/')
@login_required
def index():
    db = get_db()
    
    # Get filter parameters
    class_id = request.args.get('class')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Get date range parameters with default to current day
    today = datetime.now().strftime('%Y-%m-%d')
    from_date = request.args.get('from_date', today)
    to_date = request.args.get('to_date', today)
    
    # Get student's enrolled classes
    student_classes = db.execute('''
        SELECT DISTINCT cl.class_id, cl.class_name, c.course_code, c.course_name
        FROM student_courses sc
        JOIN courses c ON sc.course_code = c.course_code
        JOIN class_courses cc ON c.course_code = cc.course_code
        JOIN classes cl ON cc.class_id = cl.class_id
        WHERE sc.student_id = ? 
        AND sc.status = 'Active'
        AND cl.semester = sc.semester
        ORDER BY cl.class_name
    ''', [current_user.id]).fetchall()
    
    if not student_classes:
        flash('No enrolled classes found.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Base query for attendance records
    query = '''
        SELECT DISTINCT
            a.timestamp, a.status, 
            cs.date, cs.start_time, cs.end_time,
            cl.class_name, cl.class_id,
            c.course_name, c.course_code
        FROM attendance a
        JOIN class_sessions cs ON a.session_id = cs.session_id
        JOIN classes cl ON cs.class_id = cl.class_id
        JOIN class_courses cc ON cl.class_id = cc.class_id
        JOIN courses c ON cc.course_code = c.course_code
        JOIN student_courses sc ON cc.course_code = sc.course_code 
            AND sc.student_id = a.student_id
        WHERE a.student_id = ?
        AND sc.status = 'Active'
        AND cl.semester = sc.semester
        AND cs.date BETWEEN ? AND ?
    '''
    params = [current_user.id, from_date, to_date]
    
    if class_id:
        query += ' AND cl.class_id = ?'
        params.append(class_id)
    
    query += ' ORDER BY cs.date DESC, cs.start_time DESC'
    
    # Get total count for pagination
    count_query = f"SELECT COUNT(*) as total FROM ({query}) as subquery"
    total = db.execute(count_query, params).fetchone()['total']
    
    # Add pagination
    query += ' LIMIT ? OFFSET ?'
    params.extend([per_page, (page - 1) * per_page])
    
    # Get attendance records
    attendance_records = db.execute(query, params).fetchall()
    
    # Calculate overall statistics
    overall_stats = db.execute('''
        WITH total_class_count AS (
            SELECT COUNT(DISTINCT cs.session_id) as total_count
            FROM student_courses sc
            JOIN class_courses cc ON sc.course_code = cc.course_code
            JOIN classes cl ON cc.class_id = cl.class_id
            JOIN class_sessions cs ON cl.class_id = cs.class_id
            WHERE sc.student_id = ? 
            AND sc.status = 'Active'
            AND cl.semester = sc.semester
            AND cs.date <= date('now')
        ),
        attended_class_count AS (
            SELECT COUNT(DISTINCT cs.session_id) as present_count
            FROM student_courses sc
            JOIN class_courses cc ON sc.course_code = cc.course_code
            JOIN classes cl ON cc.class_id = cl.class_id
            JOIN class_sessions cs ON cl.class_id = cs.class_id
            JOIN attendance a ON cs.session_id = a.session_id 
                AND a.student_id = sc.student_id
            WHERE sc.student_id = ? 
            AND sc.status = 'Active'
            AND cl.semester = sc.semester
            AND cs.date <= date('now')
            AND a.status = 'Present'
        )
        SELECT 
            t.total_count as total_sessions,
            COALESCE(a.present_count, 0) as present_count,
            (t.total_count - COALESCE(a.present_count, 0)) as absent_count
        FROM total_class_count t
        LEFT JOIN attended_class_count a ON 1=1
    ''', [current_user.id, current_user.id]).fetchone()
    
    # Calculate attendance rate
    if overall_stats and overall_stats['total_sessions'] > 0:
        overall_stats = dict(overall_stats)
        overall_stats['attendance_rate'] = (
            overall_stats['present_count'] / overall_stats['total_sessions']
        ) * 100
    
    # Get monthly attendance statistics
    monthly_stats = db.execute('''
        WITH monthly_total_sessions AS (
            SELECT 
                strftime('%Y-%m', cs.date) as month,
                COUNT(DISTINCT cs.session_id) as total_sessions
            FROM student_courses sc
            JOIN class_courses cc ON sc.course_code = cc.course_code
            JOIN classes cl ON cc.class_id = cl.class_id
            JOIN class_sessions cs ON cl.class_id = cs.class_id
            WHERE sc.student_id = ? 
            AND sc.status = 'Active'
            AND cl.semester = sc.semester
            AND cs.date >= date('now', '-6 months')
            GROUP BY strftime('%Y-%m', cs.date)
        ),
        monthly_present_sessions AS (
            SELECT 
                strftime('%Y-%m', cs.date) as month,
                COUNT(DISTINCT cs.session_id) as present_count
            FROM student_courses sc
            JOIN class_courses cc ON sc.course_code = cc.course_code
            JOIN classes cl ON cc.class_id = cl.class_id
            JOIN class_sessions cs ON cl.class_id = cs.class_id
            JOIN attendance a ON cs.session_id = a.session_id 
                AND a.student_id = sc.student_id
            WHERE sc.student_id = ? 
            AND sc.status = 'Active'
            AND cl.semester = sc.semester
            AND cs.date >= date('now', '-6 months')
            AND a.status = 'Present'
            GROUP BY strftime('%Y-%m', cs.date)
        )
        SELECT 
            ms.month,
            ms.total_sessions,
            COALESCE(ma.present_count, 0) as present_count
        FROM monthly_total_sessions ms
        LEFT JOIN monthly_present_sessions ma ON ms.month = ma.month
        ORDER BY ms.month DESC
    ''', [current_user.id, current_user.id]).fetchall()
    
    # Convert monthly stats to list of dictionaries and ensure numeric values
    monthly_stats = [{
        'month': str(row['month']),
        'total_sessions': int(row['total_sessions']),
        'present_count': int(row['present_count'])
    } for row in monthly_stats]
    
    # Get class-wise statistics
    class_stats = db.execute('''
        WITH total_class_sessions AS (
            SELECT 
                cl.class_id,
                cl.class_name,
                c.course_name,
                COUNT(DISTINCT cs.session_id) as total_sessions
            FROM student_courses sc
            JOIN class_courses cc ON sc.course_code = cc.course_code
            JOIN classes cl ON cc.class_id = cl.class_id
            JOIN courses c ON sc.course_code = c.course_code
            JOIN class_sessions cs ON cl.class_id = cs.class_id
            WHERE sc.student_id = ? 
            AND sc.status = 'Active'
            AND cl.semester = sc.semester
            AND cs.date <= date('now')
            GROUP BY cl.class_id, cl.class_name, c.course_name
        ),
        present_class_sessions AS (
            SELECT 
                cl.class_id,
                COUNT(DISTINCT cs.session_id) as present_count
            FROM student_courses sc
            JOIN class_courses cc ON sc.course_code = cc.course_code
            JOIN classes cl ON cc.class_id = cl.class_id
            JOIN class_sessions cs ON cl.class_id = cs.class_id
            JOIN attendance a ON cs.session_id = a.session_id 
                AND a.student_id = sc.student_id
            WHERE sc.student_id = ? 
            AND sc.status = 'Active'
            AND cl.semester = sc.semester
            AND cs.date <= date('now')
            AND a.status = 'Present'
            GROUP BY cl.class_id
        )
        SELECT 
            tcs.class_id,
            tcs.class_name,
            tcs.course_name,
            tcs.total_sessions,
            COALESCE(pcs.present_count, 0) as present_count
        FROM total_class_sessions tcs
        LEFT JOIN present_class_sessions pcs ON tcs.class_id = pcs.class_id
        ORDER BY tcs.class_name
    ''', [current_user.id, current_user.id]).fetchall()
    
    # Convert class stats to list of dictionaries and ensure numeric values
    class_stats = [{
        'class_id': str(row['class_id']),
        'class_name': str(row['class_name']),
        'course_name': str(row['course_name']),
        'total_sessions': int(row['total_sessions']),
        'present_count': int(row['present_count'])
    } for row in class_stats]
    
    # Convert overall stats to dictionary if it exists
    if overall_stats:
        overall_stats = dict(overall_stats)
    
    # Convert attendance records to list of dictionaries
    attendance_records = [dict(record) for record in attendance_records]
    
    # Convert student classes to list of dictionaries
    student_classes = [dict(cls) for cls in student_classes]
    
    # Prepare pagination data
    pagination = {
        'page': page,
        'pages': (total + per_page - 1) // per_page,
    }
    
    return render_template('attendance/index.html',
                         attendance_records=attendance_records,
                         classes=student_classes,
                         overall_stats=overall_stats,
                         monthly_stats=monthly_stats,
                         class_stats=class_stats,
                         pagination=pagination,
                         from_date=from_date,
                         to_date=to_date)

@bp.route('/download')
@login_required
def download_report():
    db = get_db()
    class_id = request.args.get('class')
    format = request.args.get('format', 'excel')  # Default to excel
    
    # Get student info and enrolled classes
    student_info = db.execute('''
        SELECT s.fname, s.lname, s.student_id
        FROM students s
        WHERE s.student_id = ?
    ''', [current_user.id]).fetchone()
    
    if not student_info:
        flash('Student information not found.', 'warning')
        return redirect(url_for('attendance.index'))
    
    # Build query for attendance records
    query = '''
        SELECT 
            cs.date,
            cl.class_name,
            c.course_name,
            c.course_code,
            cs.start_time,
            cs.end_time,
            a.status,
            a.timestamp
        FROM attendance a
        JOIN class_sessions cs ON a.session_id = cs.session_id
        JOIN classes cl ON cs.class_id = cl.class_id
        JOIN courses c ON cl.course_code = c.course_code
        JOIN class_courses cc ON cl.class_id = cc.class_id
        JOIN student_courses sc ON cc.course_code = sc.course_code
        WHERE a.student_id = ? AND sc.status = 'Active'
    '''
    params = [current_user.id]
    
    if class_id:
        query += ' AND cl.class_id = ?'
        params.append(class_id)
    
    query += ' ORDER BY cs.date DESC, cs.start_time DESC'
    records = db.execute(query, params).fetchall()
    
    # Get monthly statistics
    monthly_stats = db.execute('''
        SELECT 
            strftime('%Y-%m', cs.date) as month,
            COUNT(DISTINCT cs.session_id) as total_sessions,
            SUM(CASE WHEN a.status = 'Present' THEN 1 ELSE 0 END) as present_count
        FROM class_sessions cs
        JOIN classes cl ON cs.class_id = cl.class_id
        JOIN class_courses cc ON cl.class_id = cc.class_id
        JOIN student_courses sc ON cc.course_code = sc.course_code
        LEFT JOIN attendance a ON cs.session_id = a.session_id 
            AND a.student_id = ?
        WHERE sc.student_id = ? AND sc.status = 'Active'
            AND cs.date >= date('now', '-6 months')
        GROUP BY strftime('%Y-%m', cs.date)
        ORDER BY month DESC
    ''', [current_user.id, current_user.id]).fetchall()

    if format == 'excel':
        return generate_excel_report(student_info, records, monthly_stats)
    elif format == 'pdf':
        return generate_pdf_report(student_info, records, monthly_stats)
    elif format == 'csv':
        return generate_csv_report(student_info, records)
    else:
        flash('Invalid format specified.', 'error')
        return redirect(url_for('attendance.index'))

def generate_excel_report(student_info, records, monthly_stats):
    """Generate Excel report with charts"""
    wb = Workbook()
    
    # Main attendance sheet
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add student information
    ws['A1'] = "Student Name:"
    ws['B1'] = f"{student_info['fname']} {student_info['lname']}"
    ws['A2'] = "Student ID:"
    ws['B2'] = student_info['student_id']
    ws['A3'] = "Report Generated:"
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Headers
    headers = ['Date', 'Class', 'Course', 'Time', 'Status', 'Marked At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row, record in enumerate(records, 6):
        ws.cell(row=row, column=1, value=record['date'])
        ws.cell(row=row, column=2, value=record['class_name'])
        ws.cell(row=row, column=3, value=record['course_name'])
        ws.cell(row=row, column=4, value=f"{record['start_time']} - {record['end_time']}")
        ws.cell(row=row, column=5, value=record['status'])
        ws.cell(row=row, column=6, value=record['timestamp'])
        
        # Apply borders to all cells in the row
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
    
    # Monthly Statistics Sheet
    ws_stats = wb.create_sheet("Monthly Statistics")
    
    # Add headers for monthly stats
    ws_stats['A1'] = "Monthly Attendance Statistics"
    ws_stats['A1'].font = Font(bold=True, size=14)
    
    headers = ['Month', 'Total Sessions', 'Present', 'Attendance Rate']
    for col, header in enumerate(headers, 1):
        cell = ws_stats.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add monthly statistics data
    for row, stat in enumerate(monthly_stats, 4):
        month = datetime.strptime(stat['month'], '%Y-%m')
        month_name = calendar.month_name[month.month]
        year = month.year
        
        ws_stats.cell(row=row, column=1, value=f"{month_name} {year}")
        ws_stats.cell(row=row, column=2, value=stat['total_sessions'])
        ws_stats.cell(row=row, column=3, value=stat['present_count'])
        
        if stat['total_sessions'] > 0:
            attendance_rate = (stat['present_count'] / stat['total_sessions']) * 100
            ws_stats.cell(row=row, column=4, value=f"{attendance_rate:.1f}%")
        
        # Apply borders
        for col in range(1, 5):
            ws_stats.cell(row=row, column=col).border = border
    
    # Create bar chart for monthly attendance
    chart = BarChart()
    chart.title = "Monthly Attendance Rate"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Attendance Rate (%)"
    
    data = Reference(ws_stats, min_col=4, min_row=3, max_row=len(monthly_stats)+3)
    cats = Reference(ws_stats, min_col=1, min_row=4, max_row=len(monthly_stats)+3)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws_stats.add_chart(chart, "F3")
    
    # Adjust column widths
    for ws in [ws, ws_stats]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    # Generate filename
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def generate_pdf_report(student_info, records, monthly_stats):
    """Generate PDF report with charts using WeasyPrint"""
    from flask import render_template
    # Prepare template data
    template_data = {
        'student_name': f"{student_info['fname']} {student_info['lname']}",
        'student_id': student_info['student_id'],
        'generated_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'records': records,
        'monthly_stats': monthly_stats
    }
    # Render HTML using a Jinja2 template (create attendance/report_pdf.html)
    html = render_template('attendance/report_pdf.html', **template_data)
    pdf = HTML(string=html).write_pdf()
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

def generate_csv_report(student_info, records):
    """Generate CSV report"""
    # Create string buffer for CSV data
    output = StringIO()
    writer = csv.writer(output)
    
    # Write student information
    writer.writerow(['Student Report'])
    writer.writerow(['Name:', f"{student_info['fname']} {student_info['lname']}"])
    writer.writerow(['Student ID:', student_info['student_id']])
    writer.writerow(['Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    writer.writerow([])  # Empty row for spacing
    
    # Write headers
    writer.writerow(['Date', 'Class', 'Course', 'Time', 'Status', 'Marked At'])
    
    # Write attendance records
    for record in records:
        writer.writerow([
            record['date'],
            record['class_name'],
            record['course_name'],
            f"{record['start_time']} - {record['end_time']}",
            record['status'],
            record['timestamp']
        ])
    
    # Generate filename
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    # Prepare response
    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode('utf-8')),
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )

@bp.after_request
def cleanup_temp_file(response):
    """Clean up temporary files after sending"""
    if hasattr(response, '_tmp_file_path'):
        try:
            os.unlink(response._tmp_file_path)
        except:
            pass
    return response